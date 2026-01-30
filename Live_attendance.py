import cv2
import face_recognition
import numpy as np
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook

# -------- LOAD TRAINED DATA --------
encodeListKnown = np.load("encodings.npy", allow_pickle=True)
classNames = np.load("names.npy", allow_pickle=True)

# -------- EXCEL SETUP --------
EXCEL_FILE = "Attendance.xlsx"

if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Date", "Time", "Image Path"])
    wb.save(EXCEL_FILE)

# -------- FACE IMAGE FOLDER --------
FACE_DIR = "CapturedFaces"
os.makedirs(FACE_DIR, exist_ok=True)

def markAttendance(name, face_img):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    today = datetime.now().strftime("%Y-%m-%d")

    # Prevent duplicate entry per day
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == name and row[1] == today:
            return

    now = datetime.now()
    timestamp = now.strftime("%Y-%m-%d_%H-%M-%S")

    img_path = f"{FACE_DIR}/{name}_{timestamp}.jpg"
    cv2.imwrite(img_path, face_img)

    ws.append([name, today, now.strftime("%H:%M:%S"), img_path])
    wb.save(EXCEL_FILE)

# -------- START CAMERA --------
cap = cv2.VideoCapture(0)

WINDOW_NAME = "Face Recognition Attendance"
cv2.namedWindow(WINDOW_NAME, cv2.WINDOW_NORMAL)

# ---- Center the window ----
screen_width = 1280
screen_height = 800
window_width = 800
window_height = 600

x = (screen_width - window_width) // 2
y = (screen_height - window_height) // 2

cv2.resizeWindow(WINDOW_NAME, window_width, window_height)
cv2.moveWindow(WINDOW_NAME, x, y)

print("🔴 Camera ON | Press Q to quit")

detected_names = set()   # track detections

while True:
    success, img = cap.read()
    if not success:
        print("❌ Camera not accessible")
        break

    imgSmall = cv2.resize(img, (0, 0), None, 0.25, 0.25)
    imgRGB = cv2.cvtColor(imgSmall, cv2.COLOR_BGR2RGB)

    facesCurFrame = face_recognition.face_locations(imgRGB)
    encodesCurFrame = face_recognition.face_encodings(imgRGB, facesCurFrame)

    for encodeFace, faceLoc in zip(encodesCurFrame, facesCurFrame):
        matches = face_recognition.compare_faces(encodeListKnown, encodeFace)
        faceDis = face_recognition.face_distance(encodeListKnown, encodeFace)

        matchIndex = np.argmin(faceDis)

        if matches[matchIndex]:
            name = classNames[matchIndex].upper()

            y1, x2, y2, x1 = [v * 4 for v in faceLoc]
            face_crop = img[y1:y2, x1:x2].copy()

            if name not in detected_names:
                detected_names.add(name)
                markAttendance(name, face_crop)

                print(f"✅ Attendance marked for {name}")

                # ---- AUTO CLOSE after detection (optional) ----
                # cap.release()
                # cv2.destroyAllWindows()
                # exit(0)

            cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
            cv2.putText(img, name, (x1, y1 - 10),
                        cv2.FONT_HERSHEY_SIMPLEX, 1,
                        (0, 255, 255), 2)

    cv2.imshow(WINDOW_NAME, img)

    key = cv2.waitKey(1) & 0xFF
    if key == ord('q'):
        print("🛑 Camera closed by user")
        break

# -------- CLEAN EXIT --------
cap.release()
cv2.destroyAllWindows()
print("✅ Camera released and windows closed")